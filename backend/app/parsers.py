import os
from typing import Tuple
from .config import get_settings


def _save_text(doc_id: int, text: str) -> str:
    settings = get_settings()
    out_path = os.path.join(settings.TEXT_DIR, f"doc_{doc_id}.txt")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(text)
    return out_path


def parse_pdf_to_text(file_path: str) -> str:
    import fitz  # PyMuPDF
    text_parts = []
    with fitz.open(file_path) as doc:
        for page in doc:
            text_parts.append(page.get_text())
    return "\n".join(text_parts)


def parse_docx_to_text(file_path: str) -> str:
    import docx
    d = docx.Document(file_path)
    return "\n".join([p.text for p in d.paragraphs])


def parse_xlsx_to_text(file_path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    text_parts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                text_parts.append("\t".join(cells))
    return "\n".join(text_parts)


def parse_txt_to_text(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def ocr_image_to_text(file_path: str) -> str:
    # Delegate to OCR via OpenAI
    from .vision_ocr import extract_text_from_image
    return extract_text_from_image(file_path)

